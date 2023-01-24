import openpyxl
import re

class Excel:
    def __init__(self, file) -> None:
        self.load_excel(file)
        self.sheet = self.woorkbook.worksheets[1]
        self.max_row = self.sheet.max_row
        self.max_column = self.sheet.max_column
        self.find_header()
        self.data = {"RM":[],"Código":[],"Descrição":[],"DN":[],"Código PDMS":[],"Código SAP":[]}
        self.check_excel()
    
    def load_excel(self, file):
        self.woorkbook: openpyxl.Workbook = openpyxl.load_workbook(file)
        return self.woorkbook

    def check_excel(self):
        default_row = 17
        default_rm = 1
        default_cod = 7
        default_descricao = 8
        rm = self.find_rm()
        while True:
            row = default_row
            dn_value = self.check_value(row, "DN")
            cod_tub_value = self.check_value(row, "Código PDMS")
            cod_sap_value = self.check_value(row, "Código SAP")
            if dn_value == "" and cod_tub_value == "" and cod_sap_value == "":
                break
            while True:
                self.data["RM"].append(rm)
                self.data["Código"].append(self.sheet["B"+str(default_cod)].value)
                self.data["Descrição"].append(self.sheet["B"+str(default_descricao)].value)
                self.data["DN"].append(dn_value)
                self.data["Código PDMS"].append(cod_tub_value)
                self.data["Código SAP"].append(cod_sap_value)
                row+=1
                dn_value = self.check_value(row, "DN")
                cod_tub_value = self.check_value(row, "Código PDMS")
                cod_sap_value = self.check_value(row, "Código SAP")
                if dn_value == "" and cod_tub_value == "" and cod_sap_value == "":
                    break
            aditionalRows = self.find_next_row(default_row)
            default_row+=aditionalRows
            default_rm+=aditionalRows
            default_cod+=aditionalRows
            default_descricao+=aditionalRows
    
    def find_next_row(self, row):
        initialRow = row
        while True:
            if row==self.max_row:
                break
            if re.sub(r'\s',"",str(self.sheet.cell(row, self.column["itemnº"]).value).lower()) == "itemnº":
                row+=2
                break
            row+=1
        return row - initialRow

    def find_header(self):
        header = 16
        self.column = {}
        y = 1
        while True:
            cell = self.sheet.cell(header, y)
            if type(cell)==openpyxl.cell.cell.MergedCell:
                coord = cell.coordinate
                for merged in self.sheet.merged_cells:
                    if merged.coord.split(":")[0]==coord or merged.coord.split(":")[1]==coord:
                        self.column[re.sub(r'\s',"",str(self.sheet.cell(merged.min_row, merged.min_col).value).lower())] = merged.min_col
                        y = merged.max_col
                        break
            if y == self.max_column:
                break
            y+=1
        print(self.column)

    def check_value(self, row, key):
        obj = {"DN":["dn","dimensões","especificação"],"Código PDMS":["códigopdms(cod_tub)"],"Código SAP":["códigosap(nm)"]}
        for k in obj[key]:
            if k in self.column:
                value = self.sheet.cell(row, self.column[k]).value
                if value is None:
                    return ""
                else:
                    return self.sheet.cell(row, self.column[k]).value
        return ""

    def find_rm(self):
        for y in range(1, self.max_column):
            if str(self.sheet.cell(1, y).value).__contains__("RM-5400.00-"):
                return self.sheet.cell(1, y).value